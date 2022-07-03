import openpyxl
from os import path

def load_excel(tmp_path):
    if path.exists(tmp_path):
        return openpyxl.load_workbook(filename=tmp_path)

wb = load_excel('Inventory.xlsx')

sheet = wb.active

def GetRow(id):
    for x in sheet:
        if x[0].value == id:
            return x[0].row

def AddProduct(product):
    sheet.append((product['id'],product['name'],product['price'],product['inventory']))
    wb.save('Inventory.xlsx')

def UpdateProduct(product):
    row = GetRow(product['id'])
    for column , value in enumerate((product['id'],product['name'],product['price'],product['inventory']), start=1):
        sheet.cell(row=row, column=column , value=value)
    wb.save('Inventory.xlsx')

def DeleteProduct(product):
    row = GetRow(product['id'])
    sheet.delete_rows(row)
    wb.save('Inventory.xlsx')

def SearchForId(id):
    for x in sheet:
        if x[0].value == int(id):
            return True
    return False

def GetMaxPrice():
    PoolPrice = 0
    Product = ''
    for x in sheet:
        if x[0].value != 'ID' and x[1].value != 'NAME' and x[2].value != 'PRICE' and x[3].value != 'INVENTORY':
            if(float(x[2].value) > PoolPrice):
                PoolPrice = float(x[2].value)
                Product = x[1].value
    return Product

def GetMinPrice():
    PoolPrice = 0
    Product = ''
    for x in sheet:
        if x[0].value != 'ID' and x[1].value != 'NAME' and x[2].value != 'PRICE' and x[3].value != 'INVENTORY':
            if(float(x[2].value) > PoolPrice and PoolPrice == 0):
                PoolPrice = float(x[2].value)
            elif(float(x[2].value) < PoolPrice):
                PoolPrice = float(x[2].value) 
                Product = x[1].value
    return Product

def GetAvarage():
    Suma = 0
    for x in sheet:
        if x[0].value != 'ID' and x[1].value != 'NAME' and x[2].value != 'PRICE' and x[3].value != 'INVENTORY':
            Suma +=  float(x[2].value)
    return round(Suma/(sheet.max_row - 1), 1)

def GetTotalPrice():
    Suma = 0
    for x in sheet:
        if x[0].value != 'ID' and x[1].value != 'NAME' and x[2].value != 'PRICE' and x[3].value != 'INVENTORY':
            Suma += float(x[2].value) *  int(x[3].value)
    return Suma

## solicitar datos
 
Action = input()
id , name , price , inventory = input().split()

tmpProducto = { 'id': int(id), 'name': name, 'price': float(price), 'inventory': int(inventory) }


if (SearchForId(tmpProducto['id']) and Action == 'AGREGAR') or (not SearchForId(tmpProducto['id']) and (Action == 'ACTUALIZAR' or Action == 'BORRAR')):
    print('ERROR')
else:
    if(Action == 'AGREGAR'):
        AddProduct(tmpProducto)
    elif(Action == 'ACTUALIZAR'):
        UpdateProduct(tmpProducto)
    elif(Action == 'BORRAR'):
        DeleteProduct(tmpProducto)

    print(GetMaxPrice(), GetMinPrice(), GetAvarage() , GetTotalPrice())

# code for remove empty rows 

# def removeEmptyRow():
#     for row in sheet.iter_rows():
#         if not all(cell.value for cell in row):
#             sheet.delete_rows(row[0].row, 1)
#             removeEmptyRow()

# print("Maximo de filas antes de eliminar:", sheet.max_row)
 
# for row in sheet:
#     removeEmptyRow()
    
# print("Maximo de filas despues de eliminar:",sheet.max_row)

# wb.save('Inventory.xlsx')